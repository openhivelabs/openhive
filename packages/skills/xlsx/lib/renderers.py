"""Workbook renderer.

Reads a JSON spec, builds an openpyxl Workbook, applies theme-aware styles,
merges, freeze panes, charts, conditional formatting, and tables.

All visual decisions come from the active Theme — never hard-code colours
or fonts here. When the spec sets a `style` name on a cell, look it up in
theme.styles. Free-form formatting goes through inline `style` dicts which
override the theme defaults for that cell.
"""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart,
)
from openpyxl.chart.series import Series
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
)
from openpyxl.styles import (
    Alignment, Border, Color, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .themes import CellStyle, Theme, rgb_hex


# ---------------------------------------------------------------------------
# style composition
# ---------------------------------------------------------------------------


def _resolve_style(theme: Theme, name_or_dict) -> CellStyle:
    """Return a CellStyle from either a theme style name (str) or an
    inline dict with the same field names. Inline dicts are merged onto
    the theme default so callers can say {"style": "header", "fill": [..]}
    via two passes: caller resolves the name, then patches the dict.
    """
    if name_or_dict is None:
        return CellStyle()
    if isinstance(name_or_dict, str):
        return theme.styles.get(name_or_dict, CellStyle())
    if isinstance(name_or_dict, dict):
        d = dict(name_or_dict)
        # tuple-coerce colour fields when caller used a JSON list
        for k in ("font_color", "fill", "border_color"):
            v = d.get(k)
            if isinstance(v, list) and len(v) == 3:
                d[k] = tuple(v)
        return CellStyle(**{k: v for k, v in d.items()
                            if k in CellStyle.__dataclass_fields__})
    return CellStyle()


def _merge_styles(*styles: CellStyle) -> CellStyle:
    """Right-most non-None wins. Used to layer inline overrides on top of
    a theme style ({"style": "header", "fill": ...})."""
    out: dict = {}
    for s in styles:
        for f in CellStyle.__dataclass_fields__:
            v = getattr(s, f)
            if v is None:
                continue
            if f == "border" and v is False:
                # only override border when explicitly True; False is the dataclass default
                continue
            out[f] = v
    return CellStyle(**out)


def _apply_cell_style(cell, style: CellStyle, theme: Theme) -> None:
    """Push a CellStyle through openpyxl's Font/Fill/Alignment/Border models."""
    if any([style.font_name, style.font_size, style.font_color is not None,
            style.bold is not None, style.italic is not None]):
        cell.font = Font(
            name=style.font_name or theme.body_font,
            size=style.font_size if style.font_size is not None else 11,
            color=Color(rgb_hex(style.font_color)) if style.font_color else Color(rgb_hex(theme.fg)),
            bold=bool(style.bold) if style.bold is not None else False,
            italic=bool(style.italic) if style.italic is not None else False,
        )
    if style.fill is not None:
        cell.fill = PatternFill(
            fill_type="solid",
            start_color=Color(rgb_hex(style.fill)),
            end_color=Color(rgb_hex(style.fill)),
        )
    if style.align_h or style.align_v or style.wrap_text:
        cell.alignment = Alignment(
            horizontal=style.align_h, vertical=style.align_v,
            wrap_text=bool(style.wrap_text) if style.wrap_text is not None else None,
        )
    if style.number_format:
        cell.number_format = style.number_format
    if style.border:
        side = Side(style="thin",
                    color=Color(rgb_hex(style.border_color or theme.grid)))
        cell.border = Border(left=side, right=side, top=side, bottom=side)


# ---------------------------------------------------------------------------
# value coercion
# ---------------------------------------------------------------------------


def _set_cell_value(cell, value) -> None:
    """Write a value to a cell, treating leading '=' as a formula and dict
    {"f": "...", "v": ...} as an explicit formula (with optional pre-cached
    value for callers who computed it themselves)."""
    if isinstance(value, dict) and "f" in value:
        cell.value = value["f"] if value["f"].startswith("=") else "=" + value["f"]
        return
    if isinstance(value, str) and value.startswith("="):
        cell.value = value
        return
    cell.value = value


# ---------------------------------------------------------------------------
# main entrypoint
# ---------------------------------------------------------------------------


def render_workbook(spec: dict, theme: Theme) -> Workbook:
    wb = Workbook()
    # openpyxl always starts with one sheet — we'll repurpose / drop it.
    wb.remove(wb.active)

    for sh in spec["sheets"]:
        _render_sheet(wb, sh, theme)

    return wb


def _render_sheet(wb: Workbook, sh: dict, theme: Theme) -> None:
    ws = wb.create_sheet(title=sh["name"])

    tab_color = sh.get("tab_color")
    if tab_color is not None:
        ws.sheet_properties.tabColor = _color_to_hex(tab_color)

    # row-major content
    rows = sh.get("rows")
    if rows:
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c)
                _set_cell_value(cell, val)

    # cell-major content (after rows so explicit cells override row-derived)
    for c in sh.get("cells") or []:
        cell = ws[c["ref"]]
        if "value" in c:
            _set_cell_value(cell, c["value"])
        style = _resolve_style(theme, c.get("style"))
        if style is not None:
            _apply_cell_style(cell, style, theme)

    # row-level style application via {"style_rows": [{"row": 1, "style": "header"}]}
    for sr in sh.get("style_rows") or []:
        row_idx = sr.get("row")
        style = _resolve_style(theme, sr.get("style"))
        if not isinstance(row_idx, int) or row_idx < 1:
            continue
        # find max col with content in this row
        max_c = ws.max_column
        for c in range(1, max_c + 1):
            _apply_cell_style(ws.cell(row=row_idx, column=c), style, theme)

    # range-level styling: [{"range": "A2:C5", "style": "input"|{...}}]
    for sg in sh.get("style_ranges") or []:
        rng = sg.get("range")
        style = _resolve_style(theme, sg.get("style"))
        if not isinstance(rng, str):
            continue
        for row in ws[rng]:
            for cell in row:
                _apply_cell_style(cell, style, theme)

    # number formats by range
    for nf in sh.get("number_formats") or []:
        for row in ws[nf["range"]]:
            for cell in row:
                cell.number_format = nf["format"]

    # merges
    for m in sh.get("merge") or []:
        ws.merge_cells(m)

    # freeze
    if sh.get("freeze"):
        ws.freeze_panes = sh["freeze"]

    # column widths
    for j, col in enumerate(sh.get("columns") or [], start=1):
        if col and col.get("width"):
            ws.column_dimensions[get_column_letter(j)].width = float(col["width"])

    # row heights
    for rh in sh.get("row_heights") or []:
        if rh.get("row") and rh.get("height"):
            ws.row_dimensions[int(rh["row"])].height = float(rh["height"])

    # native xlsx tables
    for tb in sh.get("tables") or []:
        _add_table(ws, tb)

    # conditional formatting
    for cf in sh.get("conditional") or []:
        _add_conditional(ws, cf, theme)

    # charts (added last so they layer above the data)
    for ch in sh.get("charts") or []:
        _add_chart(ws, ch, theme)


# ---------------------------------------------------------------------------
# tables
# ---------------------------------------------------------------------------


def _add_table(ws, tb: dict) -> None:
    style_name = tb.get("style", "TableStyleMedium2")
    table = Table(displayName=_table_safe_name(tb["name"]), ref=tb["range"])
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=tb.get("row_stripes", True),
        showColumnStripes=False,
    )
    ws.add_table(table)


def _table_safe_name(name: str) -> str:
    # Excel table names: letters, digits, underscore. Must not start with digit.
    out = "".join(ch if (ch.isalnum() or ch == "_") else "_" for ch in name)
    if not out or out[0].isdigit():
        out = "T_" + out
    return out


# ---------------------------------------------------------------------------
# conditional formatting
# ---------------------------------------------------------------------------


def _add_conditional(ws, cf: dict, theme: Theme) -> None:
    rng = cf["range"]
    kind = cf["kind"]
    if kind == "data_bar":
        color = _color_to_hex(cf.get("color") or theme.accent)
        rule = DataBarRule(
            start_type="min", end_type="max",
            color=color, showValue=cf.get("show_value", True),
        )
        ws.conditional_formatting.add(rng, rule)
    elif kind == "color_scale":
        # 3-stop scale: low (white) → mid (accent_soft) → high (accent)
        low = _color_to_hex(cf.get("low_color") or (255, 255, 255))
        mid = _color_to_hex(cf.get("mid_color") or theme.accent_soft)
        high = _color_to_hex(cf.get("high_color") or theme.accent)
        rule = ColorScaleRule(
            start_type="min", start_color=low,
            mid_type="percentile", mid_value=50, mid_color=mid,
            end_type="max", end_color=high,
        )
        ws.conditional_formatting.add(rng, rule)
    elif kind == "icon_set":
        rule = IconSetRule(
            icon_style=cf.get("icon_style", "3TrafficLights1"),
            type="percent", values=[0, 33, 67],
        )
        ws.conditional_formatting.add(rng, rule)
    elif kind == "cell_value":
        from openpyxl.formatting.rule import CellIsRule
        op = cf.get("op", "greaterThan")
        formula = cf.get("formula")
        if formula is None:
            return
        font = None
        fill = None
        if cf.get("font_color"):
            font = Font(color=Color(rgb_hex(_to_rgb(cf["font_color"]))))
        if cf.get("fill"):
            hx = rgb_hex(_to_rgb(cf["fill"]))
            fill = PatternFill(fill_type="solid",
                               start_color=Color(hx), end_color=Color(hx))
        rule = CellIsRule(operator=op, formula=[str(formula)],
                          font=font, fill=fill)
        ws.conditional_formatting.add(rng, rule)


# ---------------------------------------------------------------------------
# charts
# ---------------------------------------------------------------------------


def _add_chart(ws, ch: dict, theme: Theme) -> None:
    kind = ch["kind"]
    chart_cls = {
        "bar": BarChart, "column": BarChart,
        "line": LineChart, "area": AreaChart,
        "pie": PieChart, "scatter": ScatterChart,
    }[kind]
    chart = chart_cls()
    if hasattr(chart, "type"):
        chart.type = "bar" if kind == "bar" else "col"
    chart.title = ch.get("title")
    if "y_axis_title" in ch and hasattr(chart, "y_axis"):
        chart.y_axis.title = ch["y_axis_title"]
    if "x_axis_title" in ch and hasattr(chart, "x_axis"):
        chart.x_axis.title = ch["x_axis_title"]

    data_range = ch["data_range"]
    has_header = ch.get("titles_from_data", True)
    cats_col = ch.get("categories_in_first_column", True)

    # Parse data range to a Reference. By default we treat the first row
    # as series titles and the first column as categories, matching the
    # convention every reasonable spreadsheet user expects.
    from .ranges_compat import parse_range_letters as _parse
    try:
        c1, r1, c2, r2 = _parse(data_range)
    except Exception:
        # fallback — shove the whole range as data with no special handling
        ref = Reference(ws, range_string=f"{ws.title}!{data_range}")
        chart.add_data(ref, titles_from_data=has_header)
        ws.add_chart(chart, ch.get("anchor", "F2"))
        return

    if cats_col:
        cat_ref = Reference(ws, min_col=c1, min_row=r1 + (1 if has_header else 0),
                            max_col=c1, max_row=r2)
        data_ref = Reference(ws, min_col=c1 + 1, min_row=r1,
                             max_col=c2, max_row=r2)
    else:
        cat_ref = Reference(ws, min_col=c1, min_row=r1,
                            max_col=c2, max_row=r1)
        data_ref = Reference(ws, min_col=c1, min_row=r1 + (1 if has_header else 0),
                             max_col=c2, max_row=r2)

    chart.add_data(data_ref, titles_from_data=has_header)
    if cats_col:
        chart.set_categories(cat_ref)

    # palette
    palette = theme.chart_series or (theme.accent,)
    try:
        for i, ser in enumerate(chart.series):
            color_hex = rgb_hex(palette[i % len(palette)])
            from openpyxl.chart.shapes import GraphicalProperties
            from openpyxl.drawing.fill import ColorChoice
            ser.graphicalProperties = GraphicalProperties(solidFill=color_hex)
            if kind in ("line", "scatter", "area"):
                from openpyxl.drawing.line import LineProperties
                ser.graphicalProperties.line = LineProperties(solidFill=color_hex)
    except Exception:
        pass

    # pie: per-data-point colours (mirrors pptx fix)
    if kind == "pie" and chart.series:
        try:
            from openpyxl.chart.marker import DataPoint
            from openpyxl.chart.shapes import GraphicalProperties
            ser = chart.series[0]
            n_points = (r2 - r1) - (1 if has_header else 0)
            pts = []
            for j in range(max(0, n_points)):
                dp = DataPoint(idx=j)
                dp.graphicalProperties = GraphicalProperties(
                    solidFill=rgb_hex(palette[j % len(palette)])
                )
                pts.append(dp)
            ser.data_points = pts
        except Exception:
            pass

    chart.width = ch.get("width", 16)
    chart.height = ch.get("height", 9)

    ws.add_chart(chart, ch.get("anchor", "F2"))


# ---------------------------------------------------------------------------
# small utilities
# ---------------------------------------------------------------------------


def _color_to_hex(c) -> str:
    return rgb_hex(_to_rgb(c))


def _to_rgb(c) -> tuple[int, int, int]:
    if isinstance(c, str):
        s = c.lstrip("#")
        if len(s) == 6:
            return tuple(int(s[i:i + 2], 16) for i in (0, 2, 4))  # type: ignore[return-value]
    if isinstance(c, (list, tuple)) and len(c) == 3:
        return tuple(int(x) for x in c)  # type: ignore[return-value]
    return (0, 0, 0)
