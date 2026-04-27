"""Patch DSL — declarative edits for an existing xlsx workbook.

Selector grammar:

    sheet:Name                       sheet by name (case-sensitive)
    sheet:0                          sheet by zero-based index
    sheet:Name > A1                  one cell
    sheet:Name > A1:C5               a range
    sheet:Name > col:B               an entire column (for set_style/width)
    sheet:Name > row:3               an entire row

Operations (each entry of `operations`):

    {"op": "set_cell",       "target": "sheet:Name > B3", "value": 123}
    {"op": "set_cell",       "target": "sheet:Name > B3", "formula": "SUM(B1:B2)"}
    {"op": "set_range",      "target": "sheet:Name > A2:C5", "value": [[...], ...]}
    {"op": "set_range",      "target": "sheet:Name > A2:C5", "value": 0}
    {"op": "set_style",      "target": "sheet:Name > A1:C1", "style": "header"}
    {"op": "set_style",      "target": "sheet:Name > B2",    "style": {"bold": true, "fill": [219,234,254]}}
    {"op": "set_number_format","target": "sheet:Name > B2:B10", "format": "$#,##0"}
    {"op": "insert_rows",    "target": "sheet:Name", "before": 5, "count": 2}
    {"op": "delete_rows",    "target": "sheet:Name", "from": 5, "count": 2}
    {"op": "rename_sheet",   "target": "sheet:0",   "to": "Q1"}
    {"op": "set_tab_color",  "target": "sheet:Name", "color": "1D4ED8"}
    {"op": "add_sheet",      "name": "New", "after": "Summary"}
    {"op": "delete_sheet",   "target": "sheet:Name"}
    {"op": "update_chart_data","target": "sheet:Name > chart:0", "data_range": "A1:B10"}
"""
from __future__ import annotations

import re
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# selector parsing -----------------------------------------------------------


_SHEET_STEP_RE = re.compile(r"^\s*sheet\s*:\s*(?P<id>.+?)\s*$")


def parse_selector(s: str) -> tuple[str, list[str]]:
    """Returns (sheet_id, [tail_steps]).

    sheet_id is either an int-as-string ('0') or the literal sheet name.
    Tail steps are the raw strings after the first '>', untouched."""
    if not isinstance(s, str) or not s.strip():
        raise ValueError("selector must be a non-empty string")
    parts = [p.strip() for p in s.split(">")]
    head = parts[0]
    m = _SHEET_STEP_RE.match(head)
    if not m:
        raise ValueError(f"selector must start with 'sheet:Name' or 'sheet:N', got {head!r}")
    return m.group("id"), parts[1:]


def resolve_sheet(wb: Workbook, sheet_id: str):
    if sheet_id.isdigit():
        idx = int(sheet_id)
        names = wb.sheetnames
        if not (0 <= idx < len(names)):
            raise OpError(f"sheet index {idx} out of range (have {len(names)})")
        return wb[names[idx]]
    if sheet_id in wb.sheetnames:
        return wb[sheet_id]
    raise OpError(f"no sheet named {sheet_id!r}; available: {wb.sheetnames}")


# error type ----------------------------------------------------------------


class OpError(ValueError):
    pass


# style application via shared renderer helper -------------------------------


def _iter_cells(ws, target_range: str):
    """Yield every Cell in target_range. Handles both single-cell and
    range refs uniformly."""
    target = ws[target_range]
    if hasattr(target, "value"):
        yield target
        return
    for row in target:
        if hasattr(row, "value"):
            yield row
            continue
        for cell in row:
            yield cell


def _apply_style(ws, target_range: str, style_spec, theme) -> int:
    from lib.renderers import _apply_cell_style, _resolve_style
    style = _resolve_style(theme, style_spec)
    n = 0
    for cell in _iter_cells(ws, target_range):
        _apply_cell_style(cell, style, theme)
        n += 1
    return n


# ops -----------------------------------------------------------------------


def op_set_cell(wb, op: dict, theme) -> int:
    sheet_id, tail = parse_selector(op["target"])
    if not tail:
        raise OpError("set_cell needs a cell ref after 'sheet:...'")
    ws = resolve_sheet(wb, sheet_id)
    ref = tail[-1]
    cell = ws[ref]
    if "formula" in op:
        f = op["formula"]
        cell.value = f if str(f).startswith("=") else "=" + str(f)
    elif "value" in op:
        v = op["value"]
        if isinstance(v, str) and v.startswith("="):
            cell.value = v
        else:
            cell.value = v
    else:
        raise OpError("set_cell requires 'value' or 'formula'")
    if "style" in op:
        _apply_style(ws, ref, op["style"], theme)
    return 1


def op_set_range(wb, op: dict, theme) -> int:
    sheet_id, tail = parse_selector(op["target"])
    if not tail:
        raise OpError("set_range needs a range after 'sheet:...'")
    ws = resolve_sheet(wb, sheet_id)
    rng = tail[-1]
    if "value" not in op:
        raise OpError("set_range requires 'value' (scalar to broadcast or 2D array)")
    value = op["value"]
    n = 0
    if isinstance(value, list) and value and isinstance(value[0], list):
        # 2D — fill row-major. ws[rng] returns tuple-of-tuples for ranges,
        # bare Cell for single refs (which makes 2D fill nonsensical).
        target = ws[rng]
        if hasattr(target, "value"):
            target.value = value[0][0] if value and value[0] else None
            n = 1
        else:
            for r, row_cells in enumerate(target):
                if hasattr(row_cells, "value"):
                    if r < len(value) and value[r]:
                        row_cells.value = value[r][0]
                        n += 1
                    continue
                for c, cell in enumerate(row_cells):
                    if r < len(value) and c < len(value[r]):
                        cell.value = value[r][c]
                        n += 1
    else:
        for cell in _iter_cells(ws, rng):
            cell.value = value
            n += 1
    if "style" in op:
        _apply_style(ws, rng, op["style"], theme)
    return n


def op_set_style(wb, op: dict, theme) -> int:
    sheet_id, tail = parse_selector(op["target"])
    if not tail:
        raise OpError("set_style needs a cell or range after 'sheet:...'")
    ws = resolve_sheet(wb, sheet_id)
    return _apply_style(ws, tail[-1], op.get("style"), theme)


def op_set_number_format(wb, op: dict, theme) -> int:
    sheet_id, tail = parse_selector(op["target"])
    if not tail:
        raise OpError("set_number_format needs a cell or range")
    ws = resolve_sheet(wb, sheet_id)
    fmt = op.get("format")
    if not isinstance(fmt, str):
        raise OpError("set_number_format requires 'format' string")
    n = 0
    for cell in _iter_cells(ws, tail[-1]):
        cell.number_format = fmt
        n += 1
    return n


def op_insert_rows(wb, op: dict, theme) -> int:
    sheet_id, _ = parse_selector(op["target"])
    ws = resolve_sheet(wb, sheet_id)
    before = int(op.get("before", 1))
    count = int(op.get("count", 1))
    ws.insert_rows(before, amount=count)
    return count


def op_delete_rows(wb, op: dict, theme) -> int:
    sheet_id, _ = parse_selector(op["target"])
    ws = resolve_sheet(wb, sheet_id)
    start = int(op.get("from", 1))
    count = int(op.get("count", 1))
    ws.delete_rows(start, amount=count)
    return count


def op_rename_sheet(wb, op: dict, theme) -> int:
    sheet_id, _ = parse_selector(op["target"])
    ws = resolve_sheet(wb, sheet_id)
    new = op.get("to")
    if not isinstance(new, str) or not new:
        raise OpError("rename_sheet requires 'to' string")
    if new in wb.sheetnames and wb[new] is not ws:
        raise OpError(f"sheet name {new!r} already exists")
    ws.title = new
    return 1


def op_set_tab_color(wb, op: dict, theme) -> int:
    sheet_id, _ = parse_selector(op["target"])
    ws = resolve_sheet(wb, sheet_id)
    color = op.get("color")
    from lib.renderers import _color_to_hex
    ws.sheet_properties.tabColor = _color_to_hex(color)
    return 1


def op_add_sheet(wb, op: dict, theme) -> int:
    name = op.get("name")
    if not isinstance(name, str) or not name:
        raise OpError("add_sheet requires 'name'")
    if name in wb.sheetnames:
        raise OpError(f"sheet {name!r} already exists")
    after = op.get("after")
    pos = None
    if isinstance(after, str) and after in wb.sheetnames:
        pos = wb.sheetnames.index(after) + 1
    ws = wb.create_sheet(title=name, index=pos)
    # optional inline rows[][]
    rows = op.get("rows")
    if isinstance(rows, list):
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c).value = val
    return 1


def op_delete_sheet(wb, op: dict, theme) -> int:
    sheet_id, _ = parse_selector(op["target"])
    ws = resolve_sheet(wb, sheet_id)
    if len(wb.sheetnames) <= 1:
        raise OpError("cannot delete the last remaining sheet — Excel rejects empty workbooks")
    wb.remove(ws)
    return 1


def op_update_chart_data(wb, op: dict, theme) -> int:
    """Re-point an existing chart at a new data_range (and optionally a
    new title). Useful when the underlying numbers have moved or the
    period grew."""
    sheet_id, tail = parse_selector(op["target"])
    if len(tail) != 1 or not tail[0].startswith("chart:"):
        raise OpError("update_chart_data target must be 'sheet:Name > chart:K'")
    idx = int(tail[0].split(":", 1)[1])
    ws = resolve_sheet(wb, sheet_id)
    if not (0 <= idx < len(ws._charts)):
        raise OpError(f"chart index {idx} out of range (have {len(ws._charts)})")
    chart = ws._charts[idx]
    new_range = op.get("data_range")
    if not isinstance(new_range, str):
        raise OpError("update_chart_data requires 'data_range'")
    if "title" in op:
        chart.title = op["title"]
    # Rebuild data + categories the same way render_workbook does.
    from lib.ranges_compat import parse_range_letters
    from openpyxl.chart import Reference
    has_header = op.get("titles_from_data", True)
    cats_col = op.get("categories_in_first_column", True)
    c1, r1, c2, r2 = parse_range_letters(new_range)
    # wipe existing series
    chart.series = []
    if cats_col:
        cat_ref = Reference(ws, min_col=c1, min_row=r1 + (1 if has_header else 0),
                            max_col=c1, max_row=r2)
        data_ref = Reference(ws, min_col=c1 + 1, min_row=r1,
                             max_col=c2, max_row=r2)
        chart.add_data(data_ref, titles_from_data=has_header)
        chart.set_categories(cat_ref)
    else:
        cat_ref = Reference(ws, min_col=c1, min_row=r1, max_col=c2, max_row=r1)
        data_ref = Reference(ws, min_col=c1, min_row=r1 + (1 if has_header else 0),
                             max_col=c2, max_row=r2)
        chart.add_data(data_ref, titles_from_data=has_header)
    return 1


# dispatcher ----------------------------------------------------------------


DISPATCH = {
    "set_cell": op_set_cell,
    "set_range": op_set_range,
    "set_style": op_set_style,
    "set_number_format": op_set_number_format,
    "insert_rows": op_insert_rows,
    "delete_rows": op_delete_rows,
    "rename_sheet": op_rename_sheet,
    "set_tab_color": op_set_tab_color,
    "add_sheet": op_add_sheet,
    "delete_sheet": op_delete_sheet,
    "update_chart_data": op_update_chart_data,
}


def apply_patch(wb, patch: dict, theme) -> tuple[int, list[str]]:
    """Apply every op in patch['operations']. Returns (ops_applied, warnings).

    Not atomic — caller should open a fresh workbook each invocation so
    a mid-patch failure leaves the original on disk untouched."""
    ops = patch.get("operations")
    if not isinstance(ops, list):
        raise OpError("patch.operations must be an array")
    warnings: list[str] = []
    applied = 0
    for i, op in enumerate(ops):
        kind = op.get("op")
        fn = DISPATCH.get(kind)
        if fn is None:
            raise OpError(f"op[{i}]: unknown op {kind!r}")
        try:
            n = fn(wb, op, theme)
        except OpError as e:
            raise OpError(f"op[{i}] ({kind}): {e}") from None
        applied += 1
        # patch-time warnings
        if kind == "set_style" and n == 0:
            warnings.append(f"op[{i}] (set_style): matched 0 cells")
        if kind == "delete_rows" and op.get("count", 1) > 100:
            warnings.append(f"op[{i}] (delete_rows): {op['count']} rows — destructive, double-check the range")
    return applied, warnings
