#!/usr/bin/env python3
"""Convert a CSV file to an xlsx spec JSON.

Quick onramp for "I have a CSV, give me a styled xlsx" — produces a single
sheet with the header row styled and number_format=integer/percent applied
heuristically.

Usage:
    python csv_to_spec.py --in data.csv --out spec.json --sheet-name Sales
"""
from __future__ import annotations

import argparse
import csv
import json
import pathlib
import re
import sys

SKILL_ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SKILL_ROOT))
sys.path.insert(0, str(SKILL_ROOT.parent))

from _lib.output_path import resolve_out  # noqa: E402


_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d+\.\d+$")
_PCT_RE = re.compile(r"^-?\d+(\.\d+)?%$")


def _coerce(v: str):
    s = v.strip()
    if not s:
        return None
    if _PCT_RE.match(s):
        return float(s.rstrip("%")) / 100
    if _INT_RE.match(s):
        return int(s)
    if _FLOAT_RE.match(s):
        return float(s)
    return v


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--sheet-name", default="Sheet1")
    ap.add_argument("--theme", default="default")
    args = ap.parse_args()

    src = pathlib.Path(args.inp).expanduser()
    if not src.exists():
        return _fail(f"file not found: {src}")

    with src.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = [[_coerce(c) for c in row] for row in reader]

    if not rows:
        return _fail("empty CSV")

    # Heuristic: any column whose data cells are all percentages → percent format
    n_cols = max(len(r) for r in rows)
    nfmts = []
    for j in range(n_cols):
        col_vals = [r[j] for r in rows[1:] if j < len(r)]
        if col_vals and all(isinstance(v, float) and -1.5 <= v <= 1.5 for v in col_vals):
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(j + 1)
            nfmts.append({"range": f"{letter}2:{letter}{len(rows)}", "format": "0.0%"})

    spec = {
        "meta": {"theme": args.theme},
        "sheets": [{
            "name": args.sheet_name,
            "rows": rows,
            "freeze": "A2",
            "style_rows": [{"row": 1, "style": "header"}],
            "tables": [{"name": "Data", "range": _full_range(rows, n_cols),
                        "style": "TableStyleMedium2"}],
        }],
    }
    if nfmts:
        spec["sheets"][0]["number_formats"] = nfmts

    out = resolve_out(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(spec, ensure_ascii=False, indent=2),
                   encoding="utf-8")
    print(json.dumps({"ok": True, "path": str(out), "rows": len(rows)},
                     ensure_ascii=False))
    return 0


def _full_range(rows, n_cols) -> str:
    from openpyxl.utils import get_column_letter
    return f"A1:{get_column_letter(n_cols)}{len(rows)}"


def _fail(msg: str) -> int:
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
