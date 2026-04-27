#!/usr/bin/env python3
"""Reverse-engineer an existing .xlsx into a spec JSON.

Best-effort. Cell values + formulas + merges + freeze + tab colour +
column widths round-trip. Charts/conditional formatting/tables get noted
but don't fully reconstruct (they reference ranges; the spec just records
the range + kind so a rebuild keeps a similar visual). Pivot tables and
embedded images are not extracted.
"""
from __future__ import annotations

import argparse
import json
import pathlib
import sys

SKILL_ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SKILL_ROOT))
sys.path.insert(0, str(SKILL_ROOT.parent))

from _lib.output_path import resolve_out  # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--theme", default="default")
    args = ap.parse_args()

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        return _fail(f"missing dep: {e}")

    try:
        wb = load_workbook(args.inp, data_only=False)
    except Exception as e:
        return _fail(f"open failed: {e}")

    sheet_specs: list[dict] = []
    warnings: list[str] = []

    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        # trim trailing all-None rows
        while rows and all(v is None for v in rows[-1]):
            rows.pop()
        spec: dict = {"name": ws.title, "rows": rows}
        if ws.freeze_panes:
            spec["freeze"] = ws.freeze_panes
        if ws.sheet_properties.tabColor and ws.sheet_properties.tabColor.value:
            v = ws.sheet_properties.tabColor.value
            spec["tab_color"] = v[-6:].upper() if len(v) >= 6 else v
        merges = [str(m) for m in ws.merged_cells.ranges]
        if merges:
            spec["merge"] = merges
        # column widths (only those explicitly set)
        cols = []
        for j in range(1, ws.max_column + 1):
            letter = get_column_letter(j)
            d = ws.column_dimensions.get(letter)
            cols.append({"width": d.width} if d and d.width else {})
        if any(c for c in cols):
            spec["columns"] = cols
        # tables
        tables = []
        for t in ws.tables.values():
            tables.append({"name": t.displayName, "range": t.ref,
                           "style": t.tableStyleInfo.name if t.tableStyleInfo else None})
        if tables:
            spec["tables"] = tables
        # charts — record kind + best-guess data range
        charts = []
        for ch in ws._charts:
            kind = type(ch).__name__.replace("Chart", "").lower() or "column"
            kind = {"bar": "column", "pie": "pie", "line": "line",
                    "area": "area", "scatter": "scatter"}.get(kind, "column")
            charts.append({"kind": kind, "title": _chart_title(ch),
                           "data_range": "A1:B2",  # placeholder — real range lost
                           "anchor": "F2"})
            warnings.append(
                f"sheet[{ws.title}]: chart data_range placeholder — original "
                f"reference not recovered"
            )
        if charts:
            spec["charts"] = charts
        sheet_specs.append(spec)

    deck = {
        "meta": {"theme": args.theme, "title": ""},
        "sheets": sheet_specs,
    }

    # validate
    try:
        from lib.spec import SpecError, validate
        warnings.extend(validate(deck))
        spec_valid = True
    except SpecError as e:
        warnings.append(f"extracted spec fails build-time validate(): {e}")
        spec_valid = False

    out = resolve_out(args.out, ensure_ext=".json")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(deck, ensure_ascii=False, indent=2),
                   encoding="utf-8")
    print(json.dumps({
        "ok": True, "path": str(out),
        "sheets": len(sheet_specs),
        "spec_valid": spec_valid,
        "warnings": warnings,
    }, ensure_ascii=False))
    return 0


def _chart_title(ch) -> str | None:
    t = getattr(ch, "title", None)
    if isinstance(t, str):
        return t
    try:
        for p in t.tx.rich.p:
            for r in p.r:
                if r.t:
                    return r.t
    except Exception:
        pass
    return None


def _fail(msg: str) -> int:
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
