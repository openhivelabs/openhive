#!/usr/bin/env python3
"""Apply a patch to an existing .xlsx.

Usage:
    python edit_xlsx.py --in book.xlsx --patch patch.json --out new.xlsx
"""
from __future__ import annotations

import argparse
import json
import pathlib
import sys

SKILL_ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SKILL_ROOT))
sys.path.insert(0, str(SKILL_ROOT.parent))

from _lib.output_path import resolve_out  # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--patch", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--theme", default=None,
                    help="Theme name to use for style ops. Default: 'default'.")
    args = ap.parse_args()

    try:
        from helpers.patch import OpError, apply_patch
        from lib.themes import get_theme
        from openpyxl import load_workbook
    except ImportError as e:
        return _fail(f"missing dep: {e}")

    try:
        with open(args.patch, "r", encoding="utf-8") as f:
            patch = json.load(f)
    except Exception as e:
        return _fail(f"patch load failed: {e}")

    try:
        wb = load_workbook(args.inp)
    except Exception as e:
        return _fail(f"open failed: {e}")

    theme = get_theme(args.theme)

    try:
        applied, warnings = apply_patch(wb, patch, theme)
    except OpError as e:
        return _fail(str(e))
    except Exception as e:
        import traceback
        return _fail(f"unexpected: {e}\n{traceback.format_exc()}")

    out = resolve_out(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    print(json.dumps({
        "ok": True,
        "path": str(out),
        "sheets": len(wb.sheetnames),
        "ops_applied": applied,
        "warnings": warnings,
    }, ensure_ascii=False))
    return 0


def _fail(msg: str) -> int:
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
