#!/usr/bin/env python3
"""Rich structural summary of an .xlsx workbook.

Prints JSON with per-sheet shape, charts, tables, conditional rules,
formula count, merge ranges, freeze pane, tab colour, and the patch DSL
selectors the LLM would use to edit each piece.
"""
from __future__ import annotations

import argparse
import json
import pathlib
import sys

SKILL_ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SKILL_ROOT))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    args = ap.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        return _fail(f"missing dep: {e}")

    try:
        wb = load_workbook(args.inp, data_only=False)
    except Exception as e:
        return _fail(f"open failed: {e}")

    sheets_out = []
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        formulas = 0
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1

        cond_rules = sum(len(v) for v in ws.conditional_formatting._cf_rules.values())
        merges = [str(m) for m in ws.merged_cells.ranges]
        tab = ws.sheet_properties.tabColor.value if ws.sheet_properties.tabColor else None

        chart_details = []
        for k, ch in enumerate(ws._charts):
            chart_details.append({
                "kind": type(ch).__name__,
                "title": _chart_title(ch),
                "series": len(ch.series) if hasattr(ch, "series") else 0,
            })

        table_details = [
            {"name": t.displayName, "range": t.ref}
            for t in ws.tables.values()
        ]

        # selectors map for the LLM
        selectors = {
            "sheet": f"sheet:{name}",
            "header_row": f"sheet:{name} > A1:{_last_col_letter(ws)}1",
            "data_range": f"sheet:{name} > A2:{_last_col_letter(ws)}{ws.max_row}",
            "all": f"sheet:{name} > A1:{_last_col_letter(ws)}{ws.max_row}",
        }
        for k in range(len(ws._charts)):
            selectors[f"chart:{k}"] = f"sheet:{name} > chart:{k}"

        sheets_out.append({
            "index": i,
            "name": name,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "freeze": ws.freeze_panes,
            "tab_color": tab,
            "merges": merges,
            "formulas": formulas,
            "charts": chart_details,
            "tables": table_details,
            "conditional_rules": cond_rules,
            "selectors": selectors,
        })

    print(json.dumps({
        "ok": True,
        "sheet_count": len(sheets_out),
        "sheets": sheets_out,
    }, ensure_ascii=False))
    return 0


def _last_col_letter(ws) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(max(1, ws.max_column))


def _chart_title(ch) -> str | None:
    t = getattr(ch, "title", None)
    if t is None:
        return None
    if isinstance(t, str):
        return t
    # Title object — dig into rich text
    try:
        for p in t.tx.rich.p:
            for r in p.r:
                if r.t:
                    return r.t
    except Exception:
        pass
    return None


def _fail(msg: str) -> int:
    print(json.dumps({"ok": False, "error": msg}))
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
